import re
import os
import openpyxl
import requests
import time
import datetime
import json
import warnings
warnings.filterwarnings("ignore")
from domain import url_analysis

def remove_control_chars(s):  
    control_chars = ''.join(chr(i) for i in range(0, 32))  # 创建包含所有控制字符的字符串  
    trans_table = str.maketrans('', '', control_chars)    # 创建一个翻译表，将控制字符映射为空字符串  
    return s.translate(trans_table)  

###保存查询到的内容到表格中
def input_xlsx(have_data,path_file):
    global one_table_file
    global save_file
    #file_msg={"company":company_name,"data_type":"移动应用","which_page":0}
    which_page=0
    if one_table_file=="yes":###用一个表格文件
        url=save_file+"/"+path_file["company"]+".xlsx"
        which_page=path_file["which_page"]
    else:
        file_path = save_file+'/'+path_file["company"] # 文件夹路径
        if os.path.exists(file_path):
            pass
        else:
            print('文件夹不存在',file_path)
            os.makedirs(file_path, exist_ok=True)

        url=path_file+"/"+path_file["data_type"]+".xlsx"
    try:
        workbook=openpyxl.load_workbook(url)
        sheet_names = workbook.sheetnames

    except:
        wb = openpyxl.Workbook()
        wb.save(url)
        workbook=openpyxl.load_workbook(url)

    number_of_sheets = len(sheet_names)
    for i in range(which_page+1-number_of_sheets):
        workbook.create_sheet()

    worksheet=workbook.worksheets[which_page]
    worksheet.title = path_file["data_type"]
    hang=0
    for i in have_data:
        hang=hang+1
        lie=0
        if hang%5000==0:
            print(hang)
        for j in i:
            lie=lie+1
            if isinstance(j, list): 
                j_str=""
                for j_l in j:
                    j_str=j_str+j_l+" | "
                s_clean = remove_control_chars(j_str)
            else:
                s_clean = remove_control_chars(str(j))

            worksheet.cell(hang,lie,s_clean)
    workbook.save(filename=url)
    # print("结束保存",path_file)


####从0zone的api里获取数据
def data_api_by_0zone(next=0,query_type="site",query="",size=100):
    global my_0zone_key
    global automatic_payment
    if my_0zone_key=="":
        print("请预设0.zone的api_key，请前往：https://0.zone/plug-in-unit 获取。")

    if automatic_payment=="yes":###判断是否自动扣费
        zb_pay=1
    else:
        zb_pay=0

    headers = {'Content-Type': 'application/json'}
    url="https://0.zone/api/data/"
    time.sleep(1)
    payload=json.dumps({"query_type":query_type, "query":query, "next":next, "pagesize":size, "zone_key_id":my_0zone_key, "zb_pay": zb_pay })
    code=1
    while code!=0:
        try:
            response = requests.request("post", url, headers=headers, data=payload,verify=False).json()
            code=response["code"]
            if code!=0:
                print(response)
        except:
            print("网络请求异常")
    return response

###通过对应数据类型获取数据
def retrieve_data(keylist,query_type,query_str):
    have_data=[]
    max_num=100

    
    response=data_api_by_0zone(0,query_type,query_str,max_num)
    
    all_hits_total=int(response['total'])

    all_results = response['data']
    

    print(datetime.datetime.now(),all_hits_total,"开始：",query_type,query_str,"      \r", end="")
    while len(all_results) > 0:
        last_sort_value = response['next']

        for i in all_results:
            data_i=i
            csv_li=[]
            for k in keylist:
                key_vlue=""
                key_vlue_str=""

                if "." not in k:
                    if k in data_i:
                        key_vlue=data_i[k]
                else:
                    
                    k_0=re.split(r'[.]+',k)
                    k1=k_0[0]
                    k2=k_0[1]
                    if k1 in data_i:
                        if k2 in data_i[k1]:
                            key_vlue=data_i[k1][k2]
                if isinstance(key_vlue, list):
                    for p in key_vlue:
                        if key_vlue_str=="":
                            key_vlue_str=p
                        else:
                            key_vlue_str=key_vlue_str+" / "+p
                    if key_vlue==[]:
                        key_vlue_str=""
                if isinstance(key_vlue, dict):
                    key_vlue_str=json.dump(key_vlue)
                else:
                    key_vlue_str=key_vlue
                csv_li.append(key_vlue_str)

            if csv_li not in have_data:
                have_data.append(csv_li)
            
        print(datetime.datetime.now(),str((len(have_data)/all_hits_total)*100)[0:5]+"%","开始：",query_type,query_str,"\r", end="")
        
        if len(all_results)==max_num and all_hits_total>max_num:
            response=data_api_by_0zone(last_sort_value,query_type,query_str,max_num)
            all_results = response['data']
        else:
            all_results=[]



    return have_data



#####获取 darknet 数据
def darknet_s_mapping(company_name,new_file,app_list,domain_list,ip_list):
    
    query_type="darknet"
    have_data=[]
    have_data_duibi=[]
    keylist=[
        "title","url","tags","source","event","msg.description","msg.release_time","msg.author","detail_parsing.telegram_list"
    ]
    keylist_name=[
        "网页标题","地址","标签","来源","事件类型","网页描述（自动翻译）","内容发布时间","发布者","原文包含的TG","关键词"
    ]
    have_data.append(keylist_name)
    for i in company_name:
        query_str="(body=="+str(i)+"||description=="+str(i)+"||title=="+str(i)+")&&hot=2"
        have_data_li=retrieve_data(keylist,query_type,query_str)
        print(query_str,"\r", end="")
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)

    print("暗网：完成品牌关键词查询")
    for i in domain_list:
        if "0.zone" not in i:
            query_str="(body=="+str(i)+"||description=="+str(i)+"||title=="+str(i)+")&&hot=2"
            
            have_data_li=retrieve_data(keylist,query_type,query_str)
            print(query_str,"\r", end="")
            for h in have_data_li:
                input_i=h
                if h not in have_data_duibi:
                    have_data_duibi.append(i)
                    input_i.append(i)
                    if input_i not in have_data:
                        have_data.append(input_i)
    print("暗网：完成域名关键词查询")

    for i in app_list:
        query_str="(body=="+str(i)+"||description=="+str(i)+"||title=="+str(i)+")&&hot=2"
        have_data_li=retrieve_data(keylist,query_type,query_str)
        print(query_str,"\r", end="")
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)
    print("暗网：完成移动应用关键词查询")
                    
    for i in ip_list:
        if i!="" and  "0.zone" not in i:
            query_str="(body=="+str(i)+"||description=="+str(i)+"||title=="+str(i)+")&&hot=2"
            have_data_li=retrieve_data(keylist,query_type,query_str)
            print(query_str,"\r", end="")
            for h in have_data_li:
                input_i=h
                if h not in have_data_duibi:
                    have_data_duibi.append(i)
                    input_i.append(i)
                    if input_i not in have_data:
                        have_data.append(input_i)
    print("暗网：完成信息系统关键词查询")

    input_xlsx(have_data,new_file)
    return len(have_data)-1



#####获取代码和文档数据
def code_s_mapping(company_name,new_file,app_list,domain_list,ip_list):
    
    query_type="code"
    have_data=[]
    have_data_duibi=[]
    keylist=[
        "name","url","type","source","tags","file_extension","detail_parsing.domain_list","detail_parsing.email_list","detail_parsing.ip_list"
    ]
    keylist_name=[
        "文档","地址","类型","来源","标签","后缀","原文包含的域名","原文包含的邮箱","原文包含的IP","关键词"
    ]
    have_data.append(keylist_name)
    day_30=datetime.datetime.now()-datetime.timedelta(days=30)
    day_30=str(day_30)[0:10]
    for i in company_name:
        query_str="(code_detail=="+str(i)+"||repository.description=="+str(i)+")&&tags==!爬虫&&timestamp>="+day_30
        have_data_li=retrieve_data(keylist,query_type,query_str)
        print(query_str,"\r", end="")
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)
    print("代码：完成品牌关键词查询")


    for i in domain_list:
        query_str="(code_detail=="+str(i)+"||repository.description=="+str(i)+")&&tags==!爬虫&&timestamp>="+day_30
        
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)
    print("代码：完成域名关键词查询")

    for i in app_list:
        query_str="(code_detail=="+str(i)+"||repository.description=="+str(i)+")&&tags==!爬虫&&timestamp>="+day_30
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)
    print("代码：完成移动应用关键词查询")
    ip_lei=0
    query_str=""
    keylist_i=[]
    for i in ip_list:
        ip_lei=ip_lei+1
        if query_str=="":
            query_str="code_detail=="+str(i)+"||repository.description=="+str(i)
        else:
            query_str=query_str+"||"+"code_detail=="+str(i)+"||repository.description=="+str(i)
        
        keylist_i.append(i)
        if ip_lei%20==0:
            query_str="("+query_str+str(i)+")&&tags==!爬虫&&timestamp>="+day_30
            have_data_li=retrieve_data(keylist,query_type,query_str)
            for h in have_data_li:
                input_i=h
                if h not in have_data_duibi:
                    have_data_duibi.append(i)
                    input_i.append(i)
                    if input_i not in have_data:
                        have_data.append(input_i)
            query_str=""
            keylist_i=[]
        
    if query_str=="":
        query_str="("+query_str+str(i)+")&&timestamp>="+day_30
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)
        query_str=""
    
    print("代码：完成信息系统关键词查询")

    input_xlsx(have_data,new_file)
    return len(have_data)-1
    

#####获取信息系统数据
def site_s_mapping(company_name,new_file,app_list,domain_list):
    ctx={
        "ip_list":[],
        "num":0
    }
    query_type="site"
    ip_list=[]
    have_data_duibi=[]
    have_data=[]
    keylist=[
        "url","ip","port","title","company","country","city","os","tags","service","status_code","device_type","explore_timestamp","timestamp"
    ]
    keylist_name=[
        "URL","IP","端口","网页标题","所属公司","国家","城市","操作系统","标签","服务","网络请求状态码","设备类型","发现时间","更新时间","关键词"
    ]
    have_data.append(keylist_name)


    for i in company_name:
        query_str="company="+str(i)
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            url_str=h[0]
            ip_str=h[1]
            input_i=h
            
            url_res=url_analysis(url_str)
            if url_res["code"]=="200":
                root_domain=url_res["root_domain"]
                if root_domain not in ip_list and root_domain!="":
                    ip_list.append(root_domain)

            if ip_str not in ip_list and len(ip_str)>6:
                ip_list.append(ip_str)

            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)

    for i in company_name:
        query_str="html_banner=="+str(i)+"||banner=="+str(i)

        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)

    for i in domain_list:
        query_str="(company="")&&url_analyzer=="+str(i)
        
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)

    for i in app_list:
        query_str="(company="")&&(html_banner=="+str(i)+"||banner=="+str(i)+")"
        have_data_li=retrieve_data(keylist,query_type,query_str)
        for h in have_data_li:
            input_i=h
            if h not in have_data_duibi:
                have_data_duibi.append(i)
                input_i.append(i)
                if input_i not in have_data:
                    have_data.append(input_i)

    input_xlsx(have_data,new_file)
    ctx["ip_list"]=ip_list
    ctx["num"]=len(have_data)-1
    return ctx
    


#####获取移动应用数据
def app_s_mapping(company_name,new_file):
    
    query_type="apk"

    have_data=[]
    keylist=[
        "title","type","company","msg.introduction","source","msg.app_url"
    ]
    keylist_name=[
        "应用名称","类型","所属公司","简介","来源","来源地址"
    ]
    have_data.append(keylist_name)
    query_str=""
    for i in company_name:
        if query_str=="":
            query_str="company="+str(i)+"||description=="+(i)
        else:
            query_str=query_str+"||company="+str(i)+"||description=="+(i)

    have_data_li=retrieve_data(keylist,query_type,query_str)
    keyword_list=[]
    for i in have_data_li:
        if i not in have_data:
            have_data.append(i)
        brand_key=i[0]
        brand_key=brand_key.replace(".", "", 999)
        brand_key=brand_key.replace("1", "", 999)
        brand_key=brand_key.replace("2", "", 999)
        brand_key=brand_key.replace("3", "", 999)
        brand_key=brand_key.replace("4", "", 999)
        brand_key=brand_key.replace("5", "", 999)
        brand_key=brand_key.replace("6", "", 999)
        brand_key=brand_key.replace("7", "", 999)
        brand_key=brand_key.replace("8", "", 999)
        brand_key=brand_key.replace("9", "", 999)
        brand_key=brand_key.replace("0", "", 999)
        brand_key=brand_key.replace("。", "", 999)
        brand_key=brand_key.replace("⊙", "", 999)
        brand_key=brand_key.replace("®", "", 999)
        brand_key=brand_key.replace("⑧", "", 999)

        if brand_key not in keyword_list:
            if "title"!=brand_key:
                keyword_list.append(brand_key)

    input_xlsx(have_data,new_file)
    return keyword_list
    


#####获取域名数据
def domain_s_mapping(company_name,new_file):
    ctx={
        "num":0,
        "domain_list":[]
    }
    query_type="domain"

    have_data=[]
    keylist=[
        "root_domain","domain","company","level","icp","title","msg.ip"
    ]
    keylist_name=[
        "根域名","域名","所属公司","级别","备案","网页标题","解析IP"
    ]
    have_data.append(keylist_name)
    query_str=""
    for i in company_name:
        if query_str=="":
            query_str="company="+str(i)
        else:
            query_str=query_str+"||company="+str(i)
    have_data_li=retrieve_data(keylist,query_type,query_str)
    
    keyword_list=[]

    for i in have_data_li:
        if i not in have_data:
            have_data.append(i)
        new_key=i[1]
        root_domain=""
        url_res=url_analysis(new_key)
        if url_res["code"]=="200":
            root_domain=url_res["root_domain"]
            if root_domain not in keyword_list:
                if "domain"!=root_domain and root_domain[-1]!=".":
                    keyword_list.append(root_domain)
                    print(root_domain)
        else:
            print("异常URL：",new_key)


    input_xlsx(have_data,new_file)
    ctx["domain_list"]=keyword_list
    ctx["num"]=len(have_data)-1
    return ctx

##获取邮箱数据
def email_s_mapping(company_name,new_file,domain_list):
    
    query_type="email"

    have_data=[]
    keylist=[
        "email","company","leakage_num","mail_domain"
    ]
    keylist_name=[
        "邮箱","所属公司","泄露次数","邮箱域"
    ]
    have_data.append(keylist_name)
    query_str=""
    for i in company_name:
        if query_str=="":
            query_str="company="+str(i)
        else:
            query_str=query_str+"||company="+str(i)


    for i in domain_list:
        if query_str=="":
            query_str="email_domain=="+str(i)
        else:
            query_str=query_str+"||email_domain=="+str(i)


    have_data_li=retrieve_data(keylist,query_type,query_str)
    
    for i in have_data_li:
        if i not in have_data:
            have_data.append(i)
    input_xlsx(have_data,new_file)
    
    return len(have_data)-1



def reda_xlsx_001(org_file):
    workbook = openpyxl.load_workbook(org_file) 
    lei_num=0
    sheet_i=0
        
    for sheet in workbook.worksheets:
        sheet_title=sheet.title
        print(sheet_title)

        first_column_values = []
        for row in sheet.iter_rows(min_row=1,values_only=True):
            first_column_values.append(list(row))
        
        worksheet=workbook[sheet_title]
        worksheet.cell(1,4,str("移动应用"))
        worksheet.cell(1,5,str("域名"))
        worksheet.cell(1,6,str("邮箱"))
        worksheet.cell(1,7,str("信息系统"))
        worksheet.cell(1,8,str("代码和文档"))
        worksheet.cell(1,9,str("暗网情报"))
        for i in first_column_values:
            if lei_num>0:
                company_name=i[0]
                old_company_name=i[1]
                company_name_list=[]
                company_name_list.append(company_name)
                if old_company_name!="-" and len(old_company_name)>1:
                    company_name_list.append(company_name)


                ##记录任务执行情况，可用统计反馈
                # workbook_now=openpyxl.load_workbook(org_file)
                file_msg={"company":company_name,"data_type":"移动应用","which_page":0}
                app_list=app_s_mapping(company_name_list,file_msg)
                worksheet.cell(lei_num+1,4,str(len(app_list)))
                print(company_name,"移动应用",str(len(app_list)))
                workbook.save(filename=org_file)

                
                file_msg={"company":company_name,"data_type":"域名","which_page":1}
                domain_res=domain_s_mapping(company_name_list,file_msg)
                domain_list=domain_res["domain_list"]
                worksheet.cell(lei_num+1,5,str(domain_res["num"]))
                print(company_name,"域名",domain_res["num"])
                workbook.save(filename=org_file)


                file_msg={"company":company_name,"data_type":"邮箱","which_page":2}
                email_num=email_s_mapping(company_name_list,file_msg,domain_list)
                worksheet.cell(lei_num+1,6,str(email_num))
                print(company_name,"邮箱",email_num)
                workbook.save(filename=org_file)


                file_msg={"company":company_name,"data_type":"信息系统","which_page":3}
                ip_res=site_s_mapping(company_name_list,file_msg,app_list,domain_list)
                worksheet.cell(lei_num+1,7,str(ip_res["num"]))
                print(company_name,"信息系统",ip_res["num"])
                workbook.save(filename=org_file)
                

                file_msg={"company":company_name,"data_type":"代码和文档","which_page":4}
                code_num=code_s_mapping(company_name_list,file_msg,app_list,domain_list,ip_res["ip_list"])
                worksheet.cell(lei_num+1,8,str(code_num))
                print(company_name,"代码和文档",code_num)
                workbook.save(filename=org_file)
                
                file_msg={"company":company_name,"data_type":"暗网情报","which_page":5}
                code_num=darknet_s_mapping(company_name_list,file_msg,app_list,domain_list,ip_res["ip_list"])
                worksheet.cell(lei_num+1,9,str(code_num))
                print(company_name,"暗网情报",code_num)

                workbook.save(filename=org_file)
                print("==========================完成一个目标=============================")
            lei_num=lei_num+1
            


        sheet_i=sheet_i+1###进入下一个页签的光标

#####配置参数
org_file='公司目录.xlsx'###组织目标内容
save_file="data/测试文件夹"###生成文档保存路径
my_0zone_key='14ed9e84812fd8b8a1248bdb18e3008b'####0.zone的APIKEY
automatic_payment="yes" ####api次数不够时自动消费z币。no不消费，yes消费
one_table_file="yes" ####yes 整体数据使用一个表格文件，no，每个数据类型使用一个表格文件


reda_xlsx_001(org_file)
